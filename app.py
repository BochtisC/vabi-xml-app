import streamlit as st
import requests
import openpyxl
import json
import re
import pandas as pd
import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ===================== Στυλ κουμπιού λήψης =====================
st.markdown("""
    <style>
    .stDownloadButton button {
        background-color: #0066cc !important;
        color: white !important;
    }
    </style>
""", unsafe_allow_html=True)

# ===================== Ρυθμίσεις ClickUp =====================
API_TOKEN = "pk_82763580_PX00W04XWNJPJ2YR4M6NCNZ8WQPOLY6O"
LIST_IDS = [
    "901206264874",      # Energielabel Haaglanden
    "901511575020",      # Mijn EnergieLabel
    "901504459596",      # Energieinspectie
    "901512403352",      # Prosperos
]

LIST_MAPPINGS = {
    "901206264874": "mapping-Energielabel_Haaglanden.json",
    "901511575020": "mapping-Mijn_EnergieLabel.json",
    "901504459596": "mapping-EnergieInspectie.json",
    "901512403352": "mapping-Prosperos.json"
}

LIST_NAMES = {
    "901206264874": "Energielabel Haaglanden",
    "901511575020": "Mijn EnergieLabel",
    "901504459596": "Εnergie Inspectie",
    "901512403352": "Prosperos"
}

HEADERS = {"Authorization": API_TOKEN}

# ===================== ClickUp helpers =====================
def get_tasks(list_id):
    """Φέρνει ΟΛΑ τα tasks από μια λίστα (με όλες τις σελίδες)."""
    url = f"https://api.clickup.com/api/v2/list/{list_id}/task"
    all_tasks, page = [], 0
    while True:
        resp = requests.get(url, headers=HEADERS, params={"page": page, "archived": "false"})
        if resp.status_code != 200:
            break
        data = resp.json()
        tasks = data.get("tasks", [])
        if not tasks:
            break
        all_tasks.extend(tasks)
        if not data.get("last_page", False):
            page += 1
        else:
            break
    return all_tasks


def extract_custom_fields(task):
    """Μετατρέπει τα custom fields του ClickUp σε απλό dict, λύνει dropdown labels."""
    out, dropdown_options = {}, {}
    for f in task.get("custom_fields", []):
        if f.get("type") in ["drop_down", "dropdown"] and "type_config" in f and "options" in f["type_config"]:
            dropdown_options[f["name"]] = f["type_config"]["options"]
    for field in task.get("custom_fields", []):
        name = field.get("name"); value = field.get("value"); typ = field.get("type")
        if typ in ["drop_down", "dropdown"] and value is not None:
            options = dropdown_options.get(name, [])
            label = None
            for opt in options:
                if ("orderindex" in opt and opt["orderindex"] == value) or ("id" in opt and opt["id"] == value):
                    label = opt["name"]; break
            if label is not None and isinstance(label, str):
                label = label.strip()
            out[name] = {"id": value, "label": label if label is not None else str(value).strip()}
        elif value is not None:
            out[name] = str(value).strip() if isinstance(value, str) else str(value)
        else:
            out[name] = ""
    return out

# ===================== Excel helpers =====================
def clean_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        parts = value.strip().split()
        num = parts[0].replace(",", ".")
        return num.strip()
    return str(value).strip()

# ===================== XML helpers =====================
def update_verdiepingen_in_rekenzone(xml_text, rz_name, verdieping_values):
    """
    Αντικαθιστά το περιεχόμενο του <Verdiepingen> μόνο για την συγκεκριμένη rekenzone (με βάση το <Naam>).
    Κρατάει την δομή και αφήνει άλλα πεδία ως έχουν.
    """
    pattern = re.compile(
        rf'(<Rekenzone>.*?<Naam>{re.escape(rz_name)}</Naam>.*?<Verdiepingen[^>]*>)(.*?)(</Verdiepingen>)(.*?</Rekenzone>)',
        re.DOTALL
    )
    def replacement(match):
        head, _, tail, after = match.group(1), match.group(2), match.group(3), match.group(4)
        new_content = ""
        for val in verdieping_values:
            try:
                if float(str(val).replace(",", ".")) == 0:
                    continue
            except Exception:
                if not val or str(val).strip() == "":
                    continue
            new_content += f"<Verdieping><Gebruiksoppervlakte>{val}</Gebruiksoppervlakte></Verdieping>"
        return head + new_content + tail + after
    return pattern.sub(replacement, xml_text)


def checkmark(val):
    try:
        return f"✅ {val}" if float(str(val).replace(",", ".")) != 0 else "❌ 0"
    except:
        return val


def smart_patch_xml(xml_text, mappings, values, root_tag='Objecten'):
    """
    Εφαρμόζει τα mappings (excel/clickup/custom) στο XML με απλό regex-based patching.
    Δεν σπάμε τη δομή – απλά εισάγουμε ή ενημερώνουμε απλά tags κάτω από τον parent που δείχνει το xml_path.
    """
    xml = xml_text

    def patch_xml_tag(xml_str, xml_path, value):
        path_parts = xml_path.strip('./').split('/')
        if not path_parts:
            return xml_str
        parent_path = path_parts[:-1]
        final_tag = path_parts[-1]
        parent_tag = parent_path[-1] if parent_path else None
        new_tag = f"<{final_tag}>{value}</{final_tag}>"

        if parent_tag:
            parent_pattern = rf'(<{parent_tag}[^>]*>)(.*?)(</{parent_tag}>)'
            m = re.search(parent_pattern, xml_str, flags=re.DOTALL)
            if m:
                inside = m.group(2)
                tag_pattern = rf'(<{final_tag}>)(.*?)(</{final_tag}>)'
                if re.search(tag_pattern, inside, flags=re.DOTALL):
                    new_inside = re.sub(tag_pattern, rf'\1{value}\3', inside, flags=re.DOTALL)
                else:
                    new_inside = inside + new_tag
                return xml_str[:m.start(2)] + new_inside + xml_str[m.end(2):]
            else:
                # αν δεν βρίσκουμε τον parent, προσπαθούμε κάτω από ObjectAlgemeen
                object_algemeen_pattern = r'(<ObjectAlgemeen[^>]*>)(.*?)(</ObjectAlgemeen>)'
                obj_match = re.search(object_algemeen_pattern, xml_str, flags=re.DOTALL)
                if obj_match:
                    parent_block = f"<{parent_tag}>{new_tag}</{parent_tag}>"
                    new_content = parent_block + obj_match.group(2)
                    return xml_str[:obj_match.start(2)] + new_content + xml_str[obj_match.end(2):]
                else:
                    # αλλιώς, προσθέτουμε κάτω από το πρώτο άνοιγμα του root_tag
                    root_t = path_parts[0]
                    root_open = re.search(rf'(<{root_t}[^>]*>)', xml_str, flags=re.IGNORECASE)
                    if root_open:
                        parent_block = f"<{parent_tag}>{new_tag}</{parent_tag}>"
                        return xml_str[:root_open.end()] + parent_block + xml_str[root_open.end():]
                    else:
                        return f"<{parent_tag}>{new_tag}</{parent_tag}>" + xml_str
        else:
            root_t = path_parts[0]
            root_open = re.search(rf'(<{root_t}[^>]*>)', xml_str, flags=re.IGNORECASE)
            if root_open:
                return xml_str[:root_open.end()] + new_tag + xml_str[root_open.end():]
            else:
                return new_tag + xml_str

    for field in mappings:
        xml_path = field.get("xml_path")
        if not xml_path or "/" not in xml_path:
            continue
        value = values.get(field["field"], "")
        if isinstance(value, dict):
            xml_value_type = field.get("xml_value_type", "id")
            value = value.get(xml_value_type, value.get("id", ""))
        if isinstance(value, str):
            value = value.strip()
        if value != "":
            xml = patch_xml_tag(xml, xml_path, value)
    return xml


def extract_nonempty_fields_from_rekenzone(xml_text, rz_name, tag_list):
    """
    Επιστρέφει dict με όσα tags βρεθούν ΜΕ μη-κενό περιεχόμενο
    μέσα στο <RekenzoneAlgemeen> της συγκεκριμένης rekenzone.
    """
    result = {}
    for tag in tag_list:
        pattern = re.compile(
            rf'<Rekenzone>.*?<Naam>{re.escape(rz_name)}</Naam>.*?<RekenzoneAlgemeen>.*?<({tag})>(.*?)</\1>.*?</RekenzoneAlgemeen>',
            re.DOTALL
        )
        m = pattern.search(xml_text)
        if m and m.group(2).strip():
            result[tag] = m.group(2).strip()
    return result


def collect_base_fields_fallback(xml_text):
    """
    Παίρνει για κάθε tag την πρώτη διαθέσιμη (μη-κενή) τιμή από rz1 -> rz2 -> rz3.
    Αν δεν βρεθεί Gebruiksoppervlakte, default = '1'.
    """
    tags = ["Bouwjaar", "TypeBouwwijzeVloeren", "TypeBouwwijzeWanden", "Leidingdoorvoeren", "Gebruiksoppervlakte"]
    out = {}
    for tag in tags:
        val = ""
        for rz in ["rz1", "rz2", "rz3"]:
            got = extract_nonempty_fields_from_rekenzone(xml_text, rz, [tag]).get(tag, "")
            if got and str(got).strip():
                val = str(got).strip()
                break
        if tag == "Gebruiksoppervlakte" and (not val):
            val = "1"
        if val:
            out[tag] = val
    return out


def safe_patch_algemeen_fields(xml_text, rz_name, extra_fields):
    """
    Συμπληρώνει στο <RekenzoneAlgemeen> της συγκεκριμένης rz τα tags του extra_fields
    και βάζει το <Gebruiksoppervlakte> πριν από το <Leidingdoorvoeren>.
    - Αν tag ΔΕΝ υπάρχει → το προσθέτει.
    - Αν ΥΠΑΡΧΕΙ αλλά είναι κενό → το γεμίζει.
    - Αν ΥΠΑΡΧΕΙ και έχει τιμή → δεν το πειράζει.
    """
    pattern = re.compile(
        rf'(<Rekenzone>.*?<Naam>{re.escape(rz_name)}</Naam>.*?<RekenzoneAlgemeen>)(.*?)(</RekenzoneAlgemeen>)',
        re.DOTALL
    )

    def repl(m):
        head, content, tail = m.group(1), m.group(2), m.group(3)

        def upsert(tag, val):
            nonlocal content
            tag_re = re.compile(rf'<{tag}[^>]*>(.*?)</{tag}>', re.DOTALL)
            found = tag_re.search(content)
            if found:
                if found.group(1).strip() == "":
                    content = tag_re.sub(f"<{tag}>{val}</{tag}>", content, count=1)
            else:
                content += f"<{tag}>{val}</{tag}>"

        # 1) Όλα εκτός του GO
        for tag, val in extra_fields.items():
            if tag == "Gebruiksoppervlakte":
                continue
            upsert(tag, val)

        # 2) Ειδικός χειρισμός για GO: θέση ΠΡΙΝ το Leidingdoorvoeren
        go_val = extra_fields.get("Gebruiksoppervlakte", "1")
        go_tag = f"<Gebruiksoppervlakte>{go_val}</Gebruiksoppervlakte>"
        go_re = re.compile(r'<Gebruiksoppervlakte[^>]*>(.*?)</Gebruiksoppervlakte>', re.DOTALL)
        ldp_re = re.compile(r'(<Leidingdoorvoeren[^>]*>.*?</Leidingdoorvoeren>)', re.DOTALL)

        found_go = go_re.search(content)
        if found_go:
            # Αν υπάρχει αλλά είναι κενό → γέμισέ το
            if found_go.group(1).strip() == "":
                content = go_re.sub(go_tag, content, count=1)
        else:
            # Δεν υπάρχει: βάλε το ΠΡΙΝ το Leidingdoorvoeren, αλλιώς στο τέλος
            ldp_match = ldp_re.search(content)
            if ldp_match:
                content = content[:ldp_match.start()] + go_tag + content[ldp_match.start():]
            else:
                content += go_tag

        return head + content + tail

    return pattern.sub(repl, xml_text)

# ===================== UI =====================
st.title("XML Update")

uploaded_files = st.file_uploader(
    "Ανέβασε τα δύο αρχεία (XML + Excel με ίδιο όνομα, drag & drop μαζί)",
    type=["xml", "xlsx", "xls"],
    accept_multiple_files=True
)

if uploaded_files and len(uploaded_files) >= 2:
    # Ομαδοποίηση αρχείων κατά επέκταση & βάση ονόματος
    files_by_type = {}
    for f in uploaded_files:
        ext = f.name.rsplit('.', 1)[-1].lower()
        base = f.name.rsplit('.', 1)[0]
        files_by_type.setdefault(ext, {})[base.strip()] = f

    xml_files = files_by_type.get('xml', {})
    excel_files = {**files_by_type.get('xlsx', {}), **files_by_type.get('xls', {})}
    xml_files = {k.strip(): v for k, v in xml_files.items()}
    excel_files = {k.strip(): v for k, v in excel_files.items()}
    common_names = set(xml_files) & set(excel_files)

    if not common_names:
        st.error("Δε βρέθηκε XML & Excel με ίδιο όνομα!")
    else:
        # Επιλέγουμε το πρώτο κοινό (ή μπορείς να βάλεις selectbox)
        selected = list(common_names)[0]
        xml_file = xml_files[selected]
        excel_file = excel_files[selected]

        # Διαβάζουμε XML
        try:
            xml_text = xml_file.read().decode("utf-8")
        except Exception as e:
            st.error(f"Σφάλμα στο διάβασμα του XML: {e}")
            st.stop()

        # Διαβάζουμε Excel
        excel_file.seek(0)
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if "Algemeen" not in wb.sheetnames:
            st.error("❌ Το Excel δεν περιέχει φύλλο με όνομα 'Algemeen'!")
            st.stop()
        ws = wb["Algemeen"]

        # ------- Ανάγνωση Verdieping(en) + rz1/rz2/rz3 από A–D, γραμμές 3..13 -------
        verd_names, rz1_vals, rz2_vals, rz3_vals = [], [], [], []
        for i in range(3, 14):  # 3..13
            name = ws[f"A{i}"].value
            v1 = ws[f"B{i}"].value
            v2 = ws[f"C{i}"].value
            v3 = ws[f"D{i}"].value
            if not any([name, v1, v2, v3]):
                break

            def parse(v):
                if v is None or str(v).strip() == "":
                    return "0"
                return str(v).split()[0].replace(",", ".").strip()

            verd_names.append(str(name).strip() if name else f"Row {i}")
            rz1_vals.append(parse(v1))
            rz2_vals.append(parse(v2))
            rz3_vals.append(parse(v3))
        # ---------------------------------------------------------------------------

        # Πίνακας παρουσίασης (παραλείπει σειρές με όλα 0) — ΑΣΦΑΛΕΙΑ στο μήκος
        data = []

        def safe_float(x):
            try:
                return float(str(x).replace(",", "."))
            except:
                return 0.0

        n = min(len(verd_names), len(rz1_vals), len(rz2_vals), len(rz3_vals))
        for i in range(n):
            v1, v2, v3 = safe_float(rz1_vals[i]), safe_float(rz2_vals[i]), safe_float(rz3_vals[i])
            if v1 == 0 and v2 == 0 and v3 == 0:
                continue
            data.append({
                "Verdieping": verd_names[i],
                "rz1 (B)": checkmark(rz1_vals[i]),
                "rz2 (C)": checkmark(rz2_vals[i]),
                "rz3 (D)": checkmark(rz3_vals[i])
            })
        sum_b = sum(safe_float(x) for x in rz1_vals if safe_float(x) != 0)
        sum_c = sum(safe_float(x) for x in rz2_vals if safe_float(x) != 0)
        sum_d = sum(safe_float(x) for x in rz3_vals if safe_float(x) != 0)
        data.append({"Verdieping": "ΣΥΝΟΛΟ", "rz1 (B)": f"{sum_b:.2f}", "rz2 (C)": f"{sum_c:.2f}", "rz3 (D)": f"{sum_d:.2f}"})
        df = pd.DataFrame(data)

        st.success(f"Βρέθηκε Excel: {selected}.xlsx (φύλλο: Algemeen)")
        st.markdown("#### Τιμές Verdiepingen ανά rz (από Excel):")
        st.table(df)

        # ===================== ClickUp αναζήτηση task =====================
        clickup_status = st.empty()
        clickup_status.info(f"Ψάχνω στο ClickUp για task με όνομα: {selected}")

        task, found_list_id = None, None
        with ThreadPoolExecutor(max_workers=len(LIST_IDS)) as executor:
            future_to_list = {executor.submit(get_tasks, lid): lid for lid in LIST_IDS}
            for future in as_completed(future_to_list):
                list_id = future_to_list[future]
                try:
                    tasks = future.result()
                    clickup_status.info(f"Έλεγχος στη λίστα: {LIST_NAMES.get(list_id, list_id)}")
                    for t in tasks:
                        if t.get("name", "").strip() == selected.strip():
                            task = t
                            found_list_id = list_id
                            break
                except Exception as e:
                    st.warning(f"Σφάλμα στη λίστα {list_id}: {e}")
                if task:
                    break

        if not task:
            st.error(f"❌ Δε βρέθηκε task στο ClickUp με όνομα '{selected}'.")
            st.stop()

        mapping_file = LIST_MAPPINGS.get(found_list_id)
        if mapping_file is None:
            st.error(f"❌ Δεν βρέθηκε mapping για λίστα: {found_list_id}")
            st.stop()
        with open(mapping_file, "r", encoding="utf-8") as f:
            CLICKUP_FIELDS = json.load(f)

        # ===================== Διαβάζουμε τιμές Excel βάσει mapping =====================
        excel_values = {}
        for field in CLICKUP_FIELDS:
            if field.get("source") == "excel":
                cell = field.get("cell", "").replace(" ", "")
                value = ws[cell].value if cell else ""
                if field["field"] == "Gebruiksoppervlakte" and (value is None or value == ""):
                    value = "1"
                excel_values[field["field"]] = clean_excel_value(value)

        # Ενημερωτικό block για τιμές Excel
        excel_block = ""
        for field in CLICKUP_FIELDS:
            if field.get("source") == "excel":
                val = excel_values.get(field["field"], "")
                excel_block += f"✅ Τιμή <b>{field['ui_label']}</b> από Excel (<b>{field.get('cell','')}</b>): <b>{val}</b><br>"
        if excel_block:
            st.markdown(excel_block, unsafe_allow_html=True)

        st.success(f"Βρέθηκε task στο ClickUp: {selected} στη λίστα: {LIST_NAMES.get(found_list_id, found_list_id)}")

        # ===================== Τιμές από ClickUp task =====================
        fields = extract_custom_fields(task)

        if "date_created" in task:
            try:
                ts = int(task["date_created"]) / 1000
                fields["date_created"] = datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
            except Exception:
                fields["date_created"] = ""

        for field in CLICKUP_FIELDS:
            if field.get("source") == "clickup":
                val = fields.get(field["field"], "")
                label = val if field["field"] == "date_created" else (val["label"] if isinstance(val, dict) else val)
                icon = "✅" if label else "❌"
                st.markdown(f"{icon} <b>{field['ui_label']}</b>: <span style='color:#222'>{label or '(κενό)'}</span><br>", unsafe_allow_html=True)
            elif field.get("source") == "excel":
                val = excel_values.get(field["field"], "")
                st.markdown(f"📄 <b>{field['ui_label']}</b>: <span style='color:#222'>{val}</span>", unsafe_allow_html=True)
            elif field.get("source") == "custom":
                val = field.get("fixed_value", "")
                st.markdown(f"🔒 <b>{field['ui_label']}</b>: <span style='color:#222'>{val}</span>", unsafe_allow_html=True)

        # Συγκεντρωτικά updated_fields για mapping
        updated_fields = {}
        for field in CLICKUP_FIELDS:
            if field.get("source") == "clickup":
                val = fields.get(field["field"], "")
                if isinstance(val, dict):
                    updated_fields[field["field"]] = val
                elif isinstance(val, str):
                    updated_fields[field["field"]] = val.strip()
                else:
                    updated_fields[field["field"]] = val
            elif field.get("source") == "excel":
                updated_fields[field["field"]] = excel_values.get(field["field"], "")
            elif field.get("source") == "custom":
                updated_fields[field["field"]] = field.get("fixed_value", "")

        for k, v in updated_fields.items():
            if isinstance(v, str):
                updated_fields[k] = v.strip()

        # ===================== Επεξεργασία XML =====================
        try:
            # 1) Εφαρμογή mappings (ClickUp/Excel/Custom)
            new_xml = smart_patch_xml(xml_text, CLICKUP_FIELDS, updated_fields)

            # 2) Ενημέρωση Verdiepingen (από Excel) για κάθε rz που υπάρχει
            if re.search(r'<Rekenzone>.*?<Naam>rz1</Naam>', new_xml, re.DOTALL):
                new_xml = update_verdiepingen_in_rekenzone(new_xml, "rz1", rz1_vals)
            if re.search(r'<Rekenzone>.*?<Naam>rz2</Naam>', new_xml, re.DOTALL):
                new_xml = update_verdiepingen_in_rekenzone(new_xml, "rz2", rz2_vals)
            if re.search(r'<Rekenzone>.*?<Naam>rz3</Naam>', new_xml, re.DOTALL):
                new_xml = update_verdiepingen_in_rekenzone(new_xml, "rz3", rz3_vals)

            # 3) Fallback κληρονομιά base fields: για κάθε tag παίρνουμε πρώτη μη-κενή από rz1→rz2→rz3
            base_fields = collect_base_fields_fallback(new_xml)
            for rz_name in ["rz1", "rz2", "rz3"]:
                if re.search(rf'<Rekenzone>.*?<Naam>{rz_name}</Naam>', new_xml, re.DOTALL):
                    new_xml = safe_patch_algemeen_fields(new_xml, rz_name, base_fields)

            # 4) Έξοδος
            st.download_button(
                label="📥 Κατέβασε το νέο XML",
                data=new_xml,
                file_name=f"{selected}!.xml",
                mime="application/xml"
            )
        except Exception as e:
            st.error(f"Σφάλμα: {e}")

else:
    st.info("Ανέβασε δύο αρχεία με το ίδιο όνομα (XML & Excel) μαζί, με drag & drop.")
